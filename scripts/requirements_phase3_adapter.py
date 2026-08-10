"""Convert the licensed Requirements Annotation Phase 3 source to import CSV.

The caller supplies the externally downloaded/extracted directory. Raw review text
is used only for the documented gold join and is not written to the adapter output.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

RAW_SHA256 = "5e7f0ec51e74a8ec7e196123a3e370986396e7c5ed01fbb759f2ca3169102253"
GOLD_SHA256 = "d39fbc5357201264589ed3d749500827427498620923dc988a88c13723aee69e"
LABELS = ["feature", "none", "performance", "quality", "stability"]
LABEL_COLUMN = "which_category_best_fits_this_feedback_sentence_"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalized_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def gold_map(path: Path) -> dict[str, str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    assert sheet is not None
    rows = sheet.iter_rows(values_only=True)
    header = {str(value): index for index, value in enumerate(next(rows))}
    result: dict[str, str] = {}
    for row in rows:
        key = normalized_text(row[header["Reviews"]])
        label = normalized_text(row[header["Judgment"]]).casefold()
        if key in result and result[key] != label:
            raise ValueError("conflicting normalized gold review keys")
        result[key] = label
    workbook.close()
    return result


def convert(source_dir: Path, output: Path) -> dict[str, Any]:
    raw = source_dir / "P3-RawOutput.csv"
    gold_path = source_dir / "P3-Golden.xlsx"
    actual = {raw.name: sha256(raw), gold_path.name: sha256(gold_path)}
    expected = {raw.name: RAW_SHA256, gold_path.name: GOLD_SHA256}
    if actual != expected:
        raise ValueError(f"source checksum mismatch: {actual}")
    reference = gold_map(gold_path)
    retained: dict[tuple[str, str], dict[str, str]] = {}
    ordinary_rows = 0
    exact_duplicates = 0
    with raw.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if str(row.get("_golden") or "").upper() != "FALSE":
                continue
            ordinary_rows += 1
            item = row["_unit_id"].strip()
            worker = row["_worker_id"].strip()
            label = row[LABEL_COLUMN].strip().casefold()
            if label not in LABELS:
                raise ValueError(f"unknown source label: {label}")
            key = (item, worker)
            candidate = {
                "annotation_id": f"reqp3-{item}-{worker}",
                "item_id": item,
                "annotator_id": worker,
                "label": label,
                "event_version": "1",
                "annotation_source": "human",
                "content_ref": f"zenodo:3626185#phase3/unit/{item}",
                "gold_label": reference.get(normalized_text(row["reviews"]), ""),
                "gold_source": (
                    "benchmark_truth" if normalized_text(row["reviews"]) in reference else ""
                ),
            }
            prior = retained.get(key)
            if prior is not None:
                if prior != candidate:
                    raise ValueError(f"conflicting repeated item-worker pair: {key}")
                exact_duplicates += 1
                continue
            retained[key] = candidate
    output.parent.mkdir(parents=True, exist_ok=True)
    fields = list(next(iter(retained.values())))
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(retained[key] for key in sorted(retained))
    return {
        "adapter_version": "requirements-phase3-1.0.0",
        "source_checksums": actual,
        "ordinary_source_rows": ordinary_rows,
        "exact_duplicates_removed": exact_duplicates,
        "output_annotations": len(retained),
        "output_items": len({key[0] for key in retained}),
        "output_annotators": len({key[1] for key in retained}),
        "gold_items": len({row["item_id"] for row in retained.values() if row["gold_label"]}),
        "output_sha256": sha256(output),
        "raw_text_written": False,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    args = parser.parse_args()
    report = convert(args.source_dir, args.output)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()
